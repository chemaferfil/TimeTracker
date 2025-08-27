from flask import Blueprint, render_template, request, redirect, url_for, flash, session, send_file
from functools import wraps
from models.models import User, TimeRecord
from models.database import db
from werkzeug.security import generate_password_hash
from datetime import datetime, timedelta, date
import os
import tempfile
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

export_bp = Blueprint("export", __name__, template_folder="../templates")

# Decorator to check if user is admin
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("is_admin"):
            flash("Acceso no autorizado. Se requieren permisos de administrador.", "danger")
            return redirect(url_for("auth.login"))
        user = User.query.get(session.get("user_id"))
        if not user or not user.is_admin:
            session.clear()
            flash("Tu cuenta ya no tiene permisos de administrador.", "danger")
            return redirect(url_for("auth.login"))
        return f(*args, **kwargs)
    return decorated_function

# ========== EXPORTACIÓN PRINCIPAL CON FILTROS ==========

@export_bp.route("/excel", methods=["GET", "POST"])
@admin_required
def export_excel():
    if request.method == "POST":
        # Detectar qué botón se ha pulsado realmente (el último en el array de botones recibidos)
        botones = [k for k in request.form.keys() if k.startswith('excel_')]
        boton_pulsado = botones[-1] if botones else None

        # LOG de depuración para ver los valores recibidos
        print('--- FILTROS RECIBIDOS ---')
        print('Botones recibidos:', botones)
        print('Botón realmente pulsado:', boton_pulsado)
        print('centro1:', request.form.get('centro1'))
        print('usuario1:', request.form.get('usuario1'))
        print('centro2:', request.form.get('centro2'))
        print('categoria2:', request.form.get('categoria2'))
        print('centro3:', request.form.get('centro3'))
        print('horas3:', request.form.get('horas3'))
        print('centro4:', request.form.get('centro4'))
        print('usuario4:', request.form.get('usuario4'))
        print('categoria4:', request.form.get('categoria4'))
        print('horas4:', request.form.get('horas4'))
        print('start_date:', request.form.get('start_date'))
        print('end_date:', request.form.get('end_date'))

        # Inicializar filtros
        centro = user_id = categoria = weekly_hours = None

        # Usar los campos correctos según el botón realmente pulsado
        if boton_pulsado == "excel_centro_usuario":
            centro = request.form.get("centro1")
            user_id = request.form.get("usuario1")
        elif boton_pulsado == "excel_centro_categoria":
            centro = request.form.get("centro2")
            categoria = request.form.get("categoria2")
        elif boton_pulsado == "excel_centro_horas":
            centro = request.form.get("centro3")
            weekly_hours = request.form.get("horas3")
        elif boton_pulsado == "excel_solo_centro":
            centro = request.form.get("centro4")
        elif boton_pulsado == "excel_solo_usuario":
            user_id = request.form.get("usuario4")
        elif boton_pulsado == "excel_solo_categoria":
            categoria = request.form.get("categoria4")
        elif boton_pulsado == "excel_solo_horas":
            weekly_hours = request.form.get("horas4")
        else:
            # Compatibilidad con los filtros antiguos
            centro = request.form.get("centro")
            user_id = request.form.get("user_id")
            categoria = request.form.get("categoria")
            weekly_hours = request.form.get("weekly_hours") or request.form.get("jornada")

        start_date = request.form.get("start_date")
        end_date = request.form.get("end_date")

        # Validación fechas
        try:
            if start_date:
                start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
            else:
                start_date = date.today()
            if end_date:
                end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
            else:
                end_date = date.today()
            if end_date < start_date:
                flash("La fecha de fin no puede ser anterior a la fecha de inicio.", "danger")
                return redirect(url_for("export.export_excel"))
        except ValueError:
            flash("Formato de fecha inválido. Use YYYY-MM-DD.", "danger")
            return redirect(url_for("export.export_excel"))

        print('Valores usados para filtrar:')
        print('centro:', centro)
        print('user_id:', user_id)
        print('categoria:', categoria)
        print('weekly_hours:', weekly_hours)
        print('--------------------------')

        # JOIN explícito para evitar AmbiguousForeignKeysError
        query = TimeRecord.query.join(User, TimeRecord.user_id == User.id).filter(
            TimeRecord.date >= start_date,
            TimeRecord.date <= end_date
        )

        # El filtro de centro debe aplicarse siempre que se seleccione (antes que cualquier otro)
        if centro:
            query = query.filter(User.centro == centro)
        # El filtro de usuario solo si se selecciona uno concreto
        if user_id:
            query = query.filter(TimeRecord.user_id == user_id)
        # El filtro de categoría solo si se selecciona una concreta
        if categoria:
            query = query.filter(User.categoria == categoria)
        # El filtro de horas solo si se selecciona una concreta
        if weekly_hours:
            try:
                wh = int(weekly_hours)
                query = query.filter(User.weekly_hours.isnot(None))
                query = query.filter(db.cast(User.weekly_hours, db.Integer) == wh)
            except ValueError:
                flash("La jornada debe ser numérica.", "danger")
                return redirect(url_for("export.export_excel"))

        records = query.order_by(TimeRecord.user_id, TimeRecord.date).all()

        if not records:
            flash("No hay registros para el período y filtros seleccionados.", "warning")
            return redirect(url_for("export.export_excel"))

        # ========== GENERAR EXCEL ========== (sin cambios)
        # (Aquí va toda la lógica de generación y envío del excel como ya tienes)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registros de Fichaje"

        header = ["Usuario", "Nombre completo", "Categoría", "Centro", "Fecha", "Entrada", "Salida", "Horas Trabajadas", "Notas", "Modificado Por", "Última Actualización"]
        for col_num, header_text in enumerate(header, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header_text
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        row_num = 2
        for record in records:
            user = User.query.get(record.user_id)
            modified_by = User.query.get(record.modified_by) if record.modified_by else None

            # Calcular horas
            hours_worked = ""
            if record.check_in and record.check_out:
                time_diff = record.check_out - record.check_in
                hours = time_diff.total_seconds() / 3600
                hours_worked = f"{hours:.2f}"

            ws.cell(row=row_num, column=1).value = user.username if user else f"ID: {record.user_id}"
            ws.cell(row=row_num, column=2).value = user.full_name if user else "-"
            ws.cell(row=row_num, column=3).value = user.categoria if user and user.categoria else "-"
            ws.cell(row=row_num, column=4).value = user.centro if user and user.centro else "-"
            ws.cell(row=row_num, column=5).value = record.date.strftime("%d/%m/%Y")
            ws.cell(row=row_num, column=6).value = record.check_in.strftime("%H:%M:%S") if record.check_in else "-"
            ws.cell(row=row_num, column=7).value = record.check_out.strftime("%H:%M:%S") if record.check_out else "-"
            ws.cell(row=row_num, column=8).value = hours_worked
            ws.cell(row=row_num, column=9).value = record.notes
            ws.cell(row=row_num, column=10).value = modified_by.username if modified_by else "-"
            ws.cell(row=row_num, column=11).value = record.updated_at.strftime("%d/%m/%Y %H:%M:%S")
            row_num += 1

        for col_num, _ in enumerate(header, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 17

        fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)
        wb.save(temp_path)

        filename = f"registros_{start_date.strftime('%Y%m%d')}_a_{end_date.strftime('%Y%m%d')}.xlsx"
        return send_file(
            temp_path,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    # GET
    from routes.admin import get_admin_centro
    centro_admin = get_admin_centro()
    q = User.query.filter_by(is_active=True)
    if centro_admin:
        q = q.filter(User.centro == centro_admin)
    users = q.order_by(User.username).all()

    # Calcular lista de horas únicas y ordenadas ascendentemente para el desplegable "horas4"
    horas_sorted = sorted({u.weekly_hours for u in users if u.weekly_hours is not None})

    today = date.today().strftime('%Y-%m-%d')
    return render_template("export_excel.html", users=users, today=today, centro_admin=centro_admin, horas_sorted=horas_sorted)

# ========== EXCEL DIARIO ==========

@export_bp.route("/excel_daily")
@admin_required
def export_excel_daily():
    fecha_str = request.args.get('fecha', date.today().strftime('%Y-%m-%d'))
    try:
        fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
    except ValueError:
        flash("Formato de fecha inválido.", "danger")
        return redirect(url_for("export.export_excel"))

    records = TimeRecord.query.filter(TimeRecord.date == fecha).order_by(TimeRecord.user_id).all()
    if not records:
        flash("No hay registros para ese día.", "warning")
        return redirect(url_for("export.export_excel"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros diarios"
    header = ["Usuario", "Nombre completo", "Categoría", "Centro", "Fecha", "Entrada", "Salida", "Horas Trabajadas", "Notas"]
    for col_num, header_text in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header_text
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    for record in records:
        user = User.query.get(record.user_id)
        hours_worked = ""
        if record.check_in and record.check_out:
            time_diff = record.check_out - record.check_in
            hours = time_diff.total_seconds() / 3600
            hours_worked = f"{hours:.2f}"

        ws.cell(row=row_num, column=1).value = user.username if user else f"ID: {record.user_id}"
        ws.cell(row=row_num, column=2).value = user.full_name if user else "-"
        ws.cell(row=row_num, column=3).value = user.categoria if user and user.categoria else "-"
        ws.cell(row=row_num, column=4).value = user.centro if user and user.centro else "-"
        ws.cell(row=row_num, column=5).value = record.date.strftime("%d/%m/%Y")
        ws.cell(row=row_num, column=6).value = record.check_in.strftime("%H:%M:%S") if record.check_in else "-"
        ws.cell(row=row_num, column=7).value = record.check_out.strftime("%H:%M:%S") if record.check_out else "-"
        ws.cell(row=row_num, column=8).value = hours_worked
        ws.cell(row=row_num, column=9).value = record.notes
        row_num += 1

    for col_num, _ in enumerate(header, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 17

    fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    wb.save(temp_path)

    filename = f"registros_{fecha.strftime('%d%m%Y')}.xlsx"
    return send_file(
        temp_path,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ========== PDF DIARIO ==========

from fpdf import FPDF

@export_bp.route("/pdf_daily")
@admin_required
def export_pdf_daily():
    fecha_str = request.args.get('fecha', date.today().strftime('%Y-%m-%d'))
    try:
        fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
    except ValueError:
        flash("Formato de fecha inválido.", "danger")
        return redirect(url_for("export.export_excel"))

    records = TimeRecord.query.filter(TimeRecord.date == fecha).order_by(TimeRecord.user_id).all()
    if not records:
        flash("No hay registros para ese día.", "warning")
        return redirect(url_for("export.export_excel"))

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, f"Registros de fichaje del {fecha.strftime('%d/%m/%Y')}", ln=1, align="C")

    pdf.set_font("Arial", "B", 10)
    header = ["Usuario", "Nombre completo", "Categoría", "Centro", "Entrada", "Salida", "Horas", "Notas"]
    col_widths = [30, 40, 25, 30, 22, 22, 30, 55]

    for i, col_name in enumerate(header):
        pdf.cell(col_widths[i], 8, col_name, border=1, align="C")
    pdf.ln()

    pdf.set_font("Arial", "", 9)
    for record in records:
        user = User.query.get(record.user_id)
        hours_worked = ""
        if record.check_in and record.check_out:
            time_diff = record.check_out - record.check_in
            hours = time_diff.total_seconds() / 3600
            hours_worked = f"{hours:.2f}"

        row = [
            user.username if user else f"ID: {record.user_id}",
            user.full_name if user else "-",
            user.categoria if user and user.categoria else "-",
            user.centro if user and user.centro else "-",
            record.check_in.strftime("%H:%M:%S") if record.check_in else "-",
            record.check_out.strftime("%H:%M:%S") if record.check_out else "-",
            hours_worked,
            record.notes or ""
        ]
        for i, item in enumerate(row):
            pdf.cell(col_widths[i], 8, str(item), border=1, align="C")
        pdf.ln()

    fd, temp_path = tempfile.mkstemp(suffix='.pdf')
    os.close(fd)
    pdf.output(temp_path)

    filename = f"registros_{fecha.strftime('%d%m%Y')}.pdf"
    return send_file(
        temp_path,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )
